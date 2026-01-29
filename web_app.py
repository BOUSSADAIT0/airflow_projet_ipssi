"""
Application Web Flask pour l'extraction de factures
Remplace l'interface Tkinter par une interface web moderne
"""
from flask import Flask, render_template, request, jsonify, send_file, session
from werkzeug.utils import secure_filename
import os
import re
import json
from datetime import datetime
from pathlib import Path

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'bmp', 'tiff'}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500 MB (dossier de factures peut être volumineux)
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Créer le dossier uploads s'il n'existe pas
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('data', exist_ok=True)

def allowed_file(filename):
    """Vérifie si le fichier a une extension autorisée"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    """Page d'accueil"""
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Endpoint pour uploader un fichier facture"""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Retourner les informations du fichier
        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'message': 'Fichier uploadé avec succès'
        })
    
    return jsonify({'error': 'Type de fichier non autorisé'}), 400

@app.route('/api/upload_folder', methods=['POST'])
def upload_folder():
    """
    Reçoit plusieurs fichiers (dossier choisi sur le PC de l'utilisateur),
    les enregistre dans un sous-dossier de uploads/ et retourne le chemin.
    """
    if 'files' not in request.files and 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    folder_name = request.form.get('folder_name', '')
    if not folder_name or not re.match(r'^[a-zA-Z0-9_-]+$', folder_name):
        folder_name = 'from_pc_' + datetime.now().strftime('%Y%m%d_%H%M%S')
    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    os.makedirs(target_dir, exist_ok=True)
    files = request.files.getlist('files') or request.files.getlist('file')
    uploaded = 0
    copied_names = []
    seen = {}
    for i, f in enumerate(files):
        if f.filename == '':
            continue
        if not allowed_file(f.filename):
            continue
        raw_name = secure_filename(os.path.basename(f.filename))
        if not raw_name:
            raw_name = 'fichier'
        base, ext = os.path.splitext(raw_name)
        if not ext:
            ext = '.pdf'
        unique_name = raw_name
        if unique_name in seen:
            seen[unique_name] += 1
            unique_name = base + '_' + str(seen[unique_name]) + ext
        else:
            seen[unique_name] = 0
        filepath = os.path.join(target_dir, unique_name)
        f.save(filepath)
        uploaded += 1
        copied_names.append(unique_name)
    rel_path = os.path.join(UPLOAD_FOLDER, folder_name).replace('\\', '/')
    return jsonify({
        'success': True,
        'path': rel_path,
        'uploaded': uploaded,
        'files': copied_names,
        'message': f'{uploaded} fichier(s) copié(s) depuis votre PC'
    })

@app.route('/api/process', methods=['POST'])
def process_invoice():
    """Endpoint pour traiter une facture"""
    data = request.json
    filepath = data.get('filepath')
    method = data.get('method', 'auto')  # ollama, tesseract, auto
    
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'Fichier introuvable'}), 404
    
    try:
        # Importer le service de traitement (adapté pour fonctionner sans Tkinter)
        from invoice_processor import InvoiceProcessor
        
        processor = InvoiceProcessor()
        result = processor.process_invoice(filepath, method)
        
        if result.get('success'):
            return jsonify({
                'success': True,
                'data': result.get('data', {}),
                'message': 'Facture traitée avec succès'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Erreur inconnue lors du traitement')
            }), 500
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Erreur détaillée: {error_details}")
        return jsonify({
            'success': False,
            'error': str(e),
            'details': error_details
        }), 500

@app.route('/api/batch', methods=['POST'])
def process_batch():
    """Endpoint pour traiter plusieurs factures en lot"""
    data = request.json
    filepaths = data.get('filepaths', [])
    method = data.get('method', 'auto')
    
    if not filepaths:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    try:
        from invoice_processor import InvoiceProcessor
        processor = InvoiceProcessor()
        
        results = []
        for filepath in filepaths:
            if os.path.exists(filepath):
                result = processor.process_invoice(filepath, method)
                results.append({
                    'filepath': filepath,
                    'success': result.get('success', False),
                    'data': result.get('data', {}),
                    'error': result.get('error')
                })
            else:
                results.append({
                    'filepath': filepath,
                    'success': False,
                    'error': 'Fichier introuvable'
                })
        
        return jsonify({
            'success': True,
            'results': results,
            'total': len(results),
            'processed': sum(1 for r in results if r.get('success'))
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/export/excel', methods=['POST'])
def export_to_excel():
    """Endpoint pour exporter vers Excel (optionnel: excel_path dans le body)"""
    data = request.json or {}
    invoice_data = data.get('data', {})
    excel_path = data.get('excel_path')
    
    if not invoice_data:
        return jsonify({'error': 'Aucune donnée à exporter'}), 400
    
    try:
        from invoice_processor import InvoiceProcessor
        from user_config import resolve_excel_path
        
        processor = InvoiceProcessor()
        path = excel_path or resolve_excel_path(PROJECT_ROOT)
        result = processor.export_to_excel(invoice_data, excel_path=path)
        
        if result.get('success'):
            return jsonify({
                'success': True,
                'filepath': result.get('filepath'),
                'message': result.get('message', 'Export Excel réussi')
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Erreur lors de l\'export')
            }), 500
    
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'details': traceback.format_exc()
        }), 500

@app.route('/api/history', methods=['GET'])
def get_history():
    """Récupère l'historique des traitements"""
    try:
        history_file = os.path.join('data', 'history.json')
        if os.path.exists(history_file):
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
            return jsonify({'success': True, 'history': history})
        return jsonify({'success': True, 'history': []})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/upload_excel', methods=['POST'])
def upload_excel():
    """
    Reçoit un fichier Excel (.xlsx) depuis le PC de l'utilisateur,
    l'enregistre dans data/ et retourne le chemin pour la config.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Le fichier doit être un Excel (.xlsx)'}), 400
    os.makedirs('data', exist_ok=True)
    filename = secure_filename(os.path.basename(f.filename))
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    filepath = os.path.join('data', filename)
    f.save(os.path.join(PROJECT_ROOT, filepath))
    return jsonify({
        'success': True,
        'path': filepath.replace('\\', '/'),
        'message': 'Fichier Excel enregistré sur le serveur'
    })

@app.route('/api/download_excel', methods=['GET'])
def download_excel():
    """
    Télécharge le fichier Excel généré vers le PC de l'utilisateur.
    GET /api/download_excel?path=data/factures.xlsx
    """
    path = request.args.get('path', '').strip()
    if not path or '..' in path or path.startswith('/'):
        return jsonify({'error': 'Chemin invalide'}), 400
    full_path = os.path.join(PROJECT_ROOT, path.replace('/', os.sep))
    if not os.path.isfile(full_path):
        return jsonify({'error': 'Fichier introuvable'}), 404
    try:
        return send_file(
            full_path,
            as_attachment=True,
            download_name=os.path.basename(path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel_data', methods=['GET'])
def get_excel_data():
    """
    Lit le fichier Excel configuré et retourne les lignes (données extraites).
    Permet d'afficher le tableau au chargement de la page (données du DAG ou dernier traitement).
    """
    try:
        from user_config import resolve_excel_path
        excel_path = resolve_excel_path(PROJECT_ROOT)
        print(f"[DEBUG] get_excel_data - excel_path={excel_path}, exists={os.path.isfile(excel_path)}")
        if not os.path.isfile(excel_path):
            print("[DEBUG] Fichier Excel introuvable")
            return jsonify({'success': True, 'rows': [], 'excel_path': excel_path})
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[DEBUG] openpyxl non disponible")
            return jsonify({'success': False, 'error': 'openpyxl non disponible', 'rows': []}), 500
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        print(f"[DEBUG] En-têtes Excel: {headers}")
        print(f"[DEBUG] Nombre de lignes (max_row): {ws.max_row}")
        key_map = {
            'Date Facture': 'date',
            'Numéro Facture': 'numero_facture',
            'Fournisseur': 'fournisseur',
            'SIREN': 'siren',
            'SIRET': 'siret',
            'Adresse': 'adresse_fournisseur',
            'Code Postal': 'code_postal',
            'Ville': 'ville',
            'ID Client': 'id_client',
            'Total HT': 'total_ht',
            'TVA': 'tva',
            'Total TTC': 'total_ttc',
            'Fichier PDF': 'fichier_pdf',
        }
        rows = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            d = {}
            for i, h in enumerate(headers):
                if i < len(row) and h in key_map:
                    val = row[i]
                    d[key_map[h]] = '' if val is None else str(val).strip()
            rows.append(d)
            if row_count <= 3:
                print(f"[DEBUG] Ligne {row_count}: {d}")
        wb.close()
        print(f"[DEBUG] Total lignes lues: {len(rows)}")
        rel_path = os.path.relpath(excel_path, PROJECT_ROOT).replace('\\', '/')
        return jsonify({'success': True, 'rows': rows, 'excel_path': rel_path})
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'rows': [], 'details': traceback.format_exc()}), 500

@app.route('/api/source_files', methods=['GET'])
def list_source_files():
    """
    Liste les fichiers facture (PDF, images) du dossier source (dossier configuré + sous-dossiers).
    Query: path=uploads (optionnel, sinon config).
    """
    try:
        from user_config import resolve_source_folder
        rel = request.args.get('path', '').strip().replace('\\', '/')
        if rel:
            source_folder = os.path.abspath(os.path.join(PROJECT_ROOT, rel))
        else:
            source_folder = resolve_source_folder(PROJECT_ROOT)
        if not os.path.isdir(source_folder):
            return jsonify({'success': True, 'files': [], 'folder': rel or 'uploads'})
        extensions = ('pdf', 'png', 'jpg', 'jpeg', 'bmp', 'tiff')
        files = []
        for p in Path(source_folder).rglob('*'):
            if p.is_file() and p.suffix.lower().lstrip('.') in extensions:
                try:
                    size = p.stat().st_size
                except OSError:
                    size = 0
                rel_path = os.path.relpath(p, source_folder).replace('\\', '/')
                files.append({
                    'name': p.name,
                    'path': rel_path,
                    'size': size,
                })
        files.sort(key=lambda x: x['path'].lower())
        folder_display = os.path.relpath(source_folder, PROJECT_ROOT).replace('\\', '/') if source_folder.startswith(PROJECT_ROOT) else source_folder
        return jsonify({'success': True, 'files': files, 'folder': folder_display})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'files': []}), 500


@app.route('/api/config', methods=['GET', 'POST'])
def config():
    """GET: retourne la config. POST: enregistre la config (dossier source, Excel sortie, intervalle DAG)."""
    try:
        from user_config import load_user_config, save_user_config, DEFAULT_CONFIG
        if request.method == 'GET':
            return jsonify({'success': True, 'config': load_user_config()})
        data = request.json or {}
        config = {
            'invoice_source_path': data.get('invoice_source_path', DEFAULT_CONFIG['invoice_source_path']),
            'invoice_output_excel': data.get('invoice_output_excel', DEFAULT_CONFIG['invoice_output_excel']),
            'dag_interval_hours': float(data.get('dag_interval_hours', DEFAULT_CONFIG['dag_interval_hours'])),
        }
        save_user_config(config)
        return jsonify({'success': True, 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/process_folder', methods=['POST'])
def process_folder():
    """
    Traite toutes les factures du dossier indiqué et exporte vers le fichier Excel indiqué.
    Si invoice_source_path et invoice_output_excel sont envoyés dans le body, ils sont utilisés
    (dossier affiché dans l'interface) ; sinon on lit la config enregistrée.
    """
    try:
        from user_config import load_user_config, resolve_source_folder, resolve_excel_path
        from invoice_processor import InvoiceProcessor
        
        body = request.json or {}
        method = body.get('method', 'auto')  # auto → OpenAI si clé présente, sinon semantic
        print(f"[DEBUG] process_folder appelé - method={method}, body={body}")
        
        # Utiliser les chemins envoyés par l'interface (dossier/Excel affichés) si présents
        if body.get('invoice_source_path'):
            rel = body.get('invoice_source_path', '').strip().replace('\\', '/')
            source_folder = os.path.abspath(os.path.join(PROJECT_ROOT, rel))
        else:
            source_folder = resolve_source_folder(PROJECT_ROOT)
        if body.get('invoice_output_excel'):
            rel = body.get('invoice_output_excel', '').strip().replace('\\', '/')
            excel_path = os.path.abspath(os.path.join(PROJECT_ROOT, rel))
        else:
            excel_path = resolve_excel_path(PROJECT_ROOT)
        
        print(f"[DEBUG] source_folder={source_folder}, excel_path={excel_path}")
        print(f"[DEBUG] PROJECT_ROOT={PROJECT_ROOT}")
        
        if not os.path.isdir(source_folder):
            print(f"[DEBUG] Création du dossier {source_folder}")
            os.makedirs(source_folder, exist_ok=True)
        
        extensions = ('pdf', 'png', 'jpg', 'jpeg', 'bmp', 'tiff')
        invoice_files = []
        for f in Path(source_folder).rglob('*'):
            if f.is_file() and f.suffix.lower().lstrip('.') in extensions:
                invoice_files.append(str(f))
        invoice_files.sort()
        
        print(f"[DEBUG] Fichiers trouvés: {len(invoice_files)} fichiers")
        for f in invoice_files:
            print(f"[DEBUG]   - {f}")
        
        if not invoice_files:
            print("[DEBUG] Aucun fichier facture trouvé")
            return jsonify({
                'success': True,
                'message': 'Aucun fichier facture dans le dossier',
                'total': 0,
                'processed': 0,
                'excel_path': excel_path
            })
        
        processor = InvoiceProcessor()
        results = []
        for filepath in invoice_files:
            print(f"[DEBUG] Traitement de {filepath} avec méthode {method}")
            result = processor.process_invoice(filepath, method)
            print(f"[DEBUG] Résultat pour {os.path.basename(filepath)}: success={result.get('success')}, data_keys={list(result.get('data', {}).keys()) if result.get('data') else 'None'}")
            if result.get('success'):
                data = result.get('data', {})
                print(f"[DEBUG] Données extraites: date={data.get('date')}, numero={data.get('numero_facture')}, fournisseur={data.get('fournisseur')}, total_ttc={data.get('total_ttc')}")
                excel_result = processor.export_to_excel(data, excel_path=excel_path)
                results.append({
                    'filepath': filepath,
                    'success': True,
                    'excel_exported': excel_result.get('success', False),
                    'data': data
                })
            else:
                error_msg = result.get('error', 'Erreur inconnue')
                print(f"[DEBUG] Échec pour {os.path.basename(filepath)}: {error_msg}")
                results.append({
                    'filepath': filepath,
                    'success': False,
                    'error': error_msg,
                    'data': None
                })
        
        processed_count = sum(1 for r in results if r.get('success'))
        print(f"[DEBUG] Résumé: {processed_count}/{len(results)} fichiers traités avec succès")
        response_data = {
            'success': True,
            'total': len(results),
            'processed': processed_count,
            'results': results,
            'excel_path': excel_path
        }
        print(f"[DEBUG] Réponse JSON: total={response_data['total']}, processed={response_data['processed']}, results_count={len(response_data['results'])}")
        return jsonify(response_data)
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'details': traceback.format_exc()
        }), 500

@app.route('/api/status', methods=['GET'])
def get_status():
    """Retourne le statut de l'application"""
    return jsonify({
        'status': 'running',
        'openai_available': os.environ.get('OPENAI_API_KEY') is not None,
        'ollama_available': check_ollama_available(),
        'excel_available': True
    })

def check_ollama_available():
    """Vérifie si Ollama est disponible"""
    try:
        import requests
        response = requests.get('http://localhost:11434/api/tags', timeout=2)
        return response.status_code == 200
    except:
        return False

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
