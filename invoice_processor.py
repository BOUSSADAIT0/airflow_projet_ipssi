"""
Service de traitement des factures 
"""
import os
import json
import logging
from datetime import datetime
from pathlib import Path

# Configuration du logger
logging.basicConfig(level=logging.INFO)
app_logger = logging.getLogger(__name__)

# Configuration par défaut
CLIENT_SIRET = os.environ.get('CLIENT_SIRET', '')
EXCEL_AVAILABLE = True
PDF2IMAGE_AVAILABLE = False

# Extracteur sémantique local (sans Ollama ni API)
try:
    from semantic_extractor import SemanticExtractor, to_invoice_data
    _SEMANTIC_AVAILABLE = True
except ImportError:
    _SEMANTIC_AVAILABLE = False
    SemanticExtractor = None
    to_invoice_data = None

class WebUICallbacks:
    """Callbacks pour l'interface web (remplace TkinterUICallbacks)"""
    def __init__(self):
        self.progress_data = {}
        self.status_messages = []
    
    def update_progress(self, value, max_value=100, message=""):
        """Met à jour la progression"""
        self.progress_data = {
            'value': value,
            'max_value': max_value,
            'percent': int((value / max_value * 100)) if max_value > 0 else 0,
            'message': message
        }
        app_logger.info(f"Progression: {value}/{max_value} - {message}")
    
    def update_status(self, message, status_type="info"):
        """Met à jour le statut"""
        self.status_messages.append({
            'message': message,
            'type': status_type,
            'timestamp': datetime.now().isoformat()
        })
        app_logger.info(f"Status [{status_type}]: {message}")
    
    def display_extracted_data(self, data):
        """Affiche les données extraites"""
        self.extracted_data = data

class InvoiceProcessor:
    """Processeur de factures pour l'interface web et Airflow - Version simplifiée"""
    
    def __init__(self, excel_file_path=None, client_siret=None):
        self.excel_file_path = excel_file_path or os.path.join('data', 'factures.xlsx')
        self.client_siret = client_siret or CLIENT_SIRET
        
        # Initialiser OpenAI et Ollama
        self.openai_available = self._init_openai()
        self.ollama_available, self.ollama_models, self.ollama_base_url = self._init_ollama()
        
        # Créer les callbacks web
        self.callbacks = WebUICallbacks()
        
        app_logger.info(f"InvoiceProcessor initialisé - OpenAI: {self.openai_available}, Ollama: {self.ollama_available}")
    
    def _init_openai(self):
        """Initialise OpenAI"""
        return bool(os.environ.get('OPENAI_API_KEY'))
    
    def _init_ollama(self):
        """Initialise Ollama"""
        try:
            import requests
            base_url = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
            
            # Vérifier si Ollama est disponible
            try:
                response = requests.get(f"{base_url}/api/tags", timeout=2)
                if response.status_code == 200:
                    models_data = response.json()
                    available_models = [m.get('name', '') for m in models_data.get('models', [])]
                    
                    # Prioriser certains modèles
                    preferred_models = ["llama3.2", "mistral", "deepseek-coder"]
                    models = []
                    
                    for preferred in preferred_models:
                        for model in available_models:
                            if preferred.lower() in model.lower():
                                if model not in models:
                                    models.append(model)
                                    break
                    
                    if not models and available_models:
                        models = available_models[:3]
                    
                    app_logger.info(f"Ollama disponible avec modèles: {models}")
                    return True, models, base_url
            except Exception as e:
                app_logger.warning(f"Ollama non disponible: {e}")
            
            return False, [], base_url
        except ImportError:
            app_logger.warning("Bibliothèque requests non disponible pour Ollama")
            return False, [], "http://localhost:11434"
    
    def process_invoice(self, filepath, method='auto'):
        """
        Traite une facture et retourne les données extraites
        
        Args:
            filepath: Chemin vers le fichier (PDF ou image)
            method: Méthode d'extraction ('openai', 'ollama', 'tesseract', 'semantic', 'auto')
                   - 'openai': extraction via API OpenAI (texte → JSON structuré)
                   - 'semantic': extraction 100% locale (regex + spaCy), PDF texte direct ou OCR
        
        Returns:
            dict: Résultat avec 'success', 'data', 'error'
        """
        if not os.path.exists(filepath):
            return {
                'success': False,
                'error': f'Fichier introuvable: {filepath}'
            }
        
        try:
            self.callbacks.update_progress(10, 100, "Lecture du fichier...")
            file_ext = os.path.splitext(filepath)[1].lower()
            current_filepath = filepath

            # Auto : priorité OpenAI si clé présente, sinon semantic ou ollama/tesseract
            if method == 'auto':
                if self.openai_available:
                    method = 'openai'
                elif _SEMANTIC_AVAILABLE:
                    method = 'semantic'
                else:
                    method = 'tesseract' if not self.ollama_available else 'ollama'

            # Méthode OpenAI : extraire le texte localement, puis un seul appel API pour obtenir un JSON
            if method == 'openai':
                if not self.openai_available:
                    return {
                        'success': False,
                        'error': 'OPENAI_API_KEY non configurée (variable d\'environnement ou .env)'
                    }
                extracted_text = self._get_text_for_semantic(current_filepath, file_ext)
                if not extracted_text or not extracted_text.strip():
                    return {
                        'success': False,
                        'error': 'Aucun texte extrait du fichier (PDF/image vide ou OCR impossible)'
                    }
                self.callbacks.update_progress(50, 100, "Appel OpenAI pour extraction structurée...")
                extracted_data = self._extract_with_openai(extracted_text)
                if not extracted_data:
                    return {
                        'success': False,
                        'error': 'OpenAI n\'a pas retourné de JSON valide'
                    }
                extracted_data['fichier_pdf'] = os.path.basename(filepath)
                self.callbacks.update_progress(90, 100, "Enrichissement des données...")
                if not extracted_data.get('siret'):
                    extracted_data = self._enrich_with_siret(extracted_data)
                self.callbacks.update_progress(100, 100, "Terminé!")
                return {
                    'success': True,
                    'data': extracted_data,
                    'filepath': current_filepath,
                    'method_used': 'openai'
                }

            # Méthode sémantique : extraction locale (sans Ollama ni API)
            if method == 'semantic':
                if not _SEMANTIC_AVAILABLE:
                    return {
                        'success': False,
                        'error': 'Module semantic_extractor non disponible'
                    }
                extracted_text = self._get_text_for_semantic(current_filepath, file_ext)
                if not extracted_text or not extracted_text.strip():
                    return {
                        'success': False,
                        'error': 'Aucun texte extrait du fichier'
                    }
                self.callbacks.update_progress(70, 100, "Extraction sémantique locale...")
                structured = SemanticExtractor().structure_data(extracted_text)
                extracted_data = to_invoice_data(structured)
                self.callbacks.update_progress(90, 100, "Enrichissement des données...")
                if not extracted_data.get('siret'):
                    extracted_data = self._enrich_with_siret(extracted_data)
                self.callbacks.update_progress(100, 100, "Terminé!")
                return {
                    'success': True,
                    'data': extracted_data,
                    'filepath': current_filepath,
                    'method_used': 'semantic'
                }

            # Méthodes classiques : PDF → image si besoin, puis ollama ou tesseract
            if file_ext == '.pdf':
                self.callbacks.update_progress(20, 100, "Conversion PDF en image...")
                app_logger.info(f"[DEBUG] Traitement PDF avec méthode {method}")
                image_path = self._convert_pdf_to_image(filepath)
                if not image_path:
                    error_msg = f'Impossible de convertir le PDF en image. Vérifiez que pdf2image (avec poppler) ou PyMuPDF est installé.'
                    app_logger.error(f"[DEBUG] {error_msg}")
                    return {
                        'success': False,
                        'error': error_msg
                    }
                current_filepath = image_path
                app_logger.info(f"[DEBUG] PDF converti en image: {image_path}")

            self.callbacks.update_progress(40, 100, f"Extraction avec {method}...")
            extracted_text = self._extract_text(current_filepath, method)
            if not extracted_text:
                return {
                    'success': False,
                    'error': 'Aucun texte extrait du fichier'
                }
            self.callbacks.update_progress(70, 100, "Analyse des données...")
            extracted_data = self._parse_invoice_data(extracted_text)
            self.callbacks.update_progress(90, 100, "Enrichissement des données...")
            if not extracted_data.get('siret'):
                extracted_data = self._enrich_with_siret(extracted_data)
            self.callbacks.update_progress(100, 100, "Terminé!")
            return {
                'success': True,
                'data': extracted_data,
                'filepath': current_filepath,
                'method_used': method
            }
        except Exception as e:
            app_logger.error(f"Erreur lors du traitement de {filepath}: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }
    
    def _extract_text_from_pdf(self, pdf_path, max_pages=10):
        """
        Extrait le texte directement d'un PDF (sans OCR).
        Utilise PyMuPDF (fitz) si disponible. Utile pour les PDF avec texte sélectionnable.
        """
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(pdf_path)
            parts = []
            for i in range(min(len(doc), max_pages)):
                page = doc[i]
                parts.append(page.get_text())
            doc.close()
            text = "\n".join(parts).strip()
            return text if text else None
        except ImportError:
            app_logger.debug("PyMuPDF (fitz) non installé pour extraction texte PDF")
            return None
        except Exception as e:
            app_logger.warning(f"Erreur extraction texte PDF: {e}")
            return None

    def _get_text_for_semantic(self, filepath, file_ext):
        """
        Récupère le texte pour l'extraction sémantique.
        Pour un PDF : essaie d'abord l'extraction texte directe, sinon PDF→image puis Tesseract.
        Pour une image : Tesseract OCR.
        """
        if file_ext == '.pdf':
            self.callbacks.update_progress(25, 100, "Extraction texte PDF directe...")
            text = self._extract_text_from_pdf(filepath)
            if text and len(text.strip()) > 50:
                return text
            self.callbacks.update_progress(30, 100, "Conversion PDF en image (OCR)...")
            image_path = self._convert_pdf_to_image(filepath)
            if not image_path:
                return None
            self.callbacks.update_progress(45, 100, "OCR Tesseract...")
            return self._extract_with_tesseract(image_path)
        # Image
        self.callbacks.update_progress(40, 100, "OCR Tesseract...")
        return self._extract_with_tesseract(filepath)

    # Prompt système unique : on explique une fois à l'IA le format JSON attendu.
    _OPENAI_SYSTEM_PROMPT = """Tu es un assistant qui extrait les données de factures.
À partir du texte brut d'une facture (français ou autre), extrais les champs suivants et retourne UNIQUEMENT un objet JSON valide, sans markdown ni commentaire.
Clés attendues (utilise "" si absent) :
- nom_client : nom ou raison sociale du client
- nom_fournisseur : nom ou raison sociale du fournisseur / vendeur
- date : date de la facture (format JJ/MM/AAAA ou AAAA-MM-JJ)
- numero_facture : numéro de facture
- montant_ht : montant HT (nombre, sans symbole)
- tva : montant TVA ou taux TVA (nombre)
- montant_ttc : montant TTC (nombre)
- siret : SIRET (14 chiffres)
- siren : SIREN (9 chiffres)
- adresse_fournisseur : adresse du fournisseur
- code_postal : code postal
- ville : ville
Réponds uniquement par le JSON, rien d'autre."""

    def _extract_with_openai(self, invoice_text):
        """
        Un seul appel OpenAI : on envoie le texte de la facture, l'IA retourne un JSON structuré.
        Le format de sortie est normalisé pour correspondre à export_to_excel.
        """
        api_key = os.environ.get('OPENAI_API_KEY')
        if not api_key:
            app_logger.warning("OPENAI_API_KEY manquante")
            return None
        try:
            from openai import OpenAI
            client = OpenAI(api_key=api_key)
        except ImportError:
            app_logger.warning("openai non installé: pip install openai")
            return None
        # Limiter la taille pour rester dans les limites de contexte
        text = (invoice_text[:14000] + "...") if len(invoice_text) > 14000 else invoice_text
        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": self._OPENAI_SYSTEM_PROMPT},
                    {"role": "user", "content": f"Extrais les données de cette facture et retourne uniquement le JSON.\n\n{text}"}
                ],
                temperature=0.1,
                max_tokens=1500,
            )
            raw = (response.choices[0].message.content or "").strip()
            # Retirer un éventuel bloc markdown ```json ... ```
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.lower().startswith("json"):
                    raw = raw[4:].strip()
            data = json.loads(raw)
            # Mapping vers le format attendu par export_to_excel
            return {
                "date": str(data.get("date") or data.get("date_facture") or ""),
                "numero_facture": str(data.get("numero_facture") or ""),
                "fournisseur": str(data.get("nom_fournisseur") or data.get("fournisseur") or ""),
                "siren": str(data.get("siren") or ""),
                "siret": str(data.get("siret") or ""),
                "adresse_fournisseur": str(data.get("adresse_fournisseur") or data.get("adresse") or ""),
                "code_postal": str(data.get("code_postal") or ""),
                "ville": str(data.get("ville") or ""),
                "id_client": str(data.get("nom_client") or data.get("client") or ""),
                "total_ht": str(data.get("montant_ht") or data.get("total_ht") or ""),
                "tva": str(data.get("tva") or ""),
                "total_ttc": str(data.get("montant_ttc") or data.get("total_ttc") or ""),
                "fichier_pdf": "",
            }
        except json.JSONDecodeError as e:
            app_logger.error(f"OpenAI a retourné un JSON invalide: {e}")
            return None
        except Exception as e:
            app_logger.error(f"Erreur appel OpenAI: {e}", exc_info=True)
            return None

    def _convert_pdf_to_image(self, pdf_path):
        """Convertit un PDF en image. Les images temporaires sont écrites dans /tmp/invoice_images (évite les erreurs de permission dans uploads/)."""
        app_logger.info(f"[DEBUG] Conversion PDF en image: {pdf_path}")
        tmp_dir = "/tmp/invoice_images"
        os.makedirs(tmp_dir, exist_ok=True)
        base_name = os.path.basename(pdf_path).replace('.pdf', '') + '_' + str(os.getpid()) + '_page1.png'
        image_path = os.path.join(tmp_dir, base_name)
        try:
            # Essayer pdf2image d'abord
            try:
                from pdf2image import convert_from_path
                app_logger.info("[DEBUG] Tentative avec pdf2image...")
                images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=200)
                if images:
                    images[0].save(image_path, 'PNG')
                    app_logger.info(f"[DEBUG] ✓ Conversion réussie avec pdf2image: {image_path}")
                    return image_path
                else:
                    app_logger.warning("[DEBUG] pdf2image n'a retourné aucune image")
            except ImportError as e:
                app_logger.warning(f"[DEBUG] pdf2image non disponible: {e}")
            except Exception as e:
                app_logger.error(f"[DEBUG] Erreur avec pdf2image: {e}", exc_info=True)
            
            # Fallback: utiliser PyMuPDF (fitz)
            try:
                import fitz  # PyMuPDF
                app_logger.info("[DEBUG] Tentative avec PyMuPDF (fitz)...")
                doc = fitz.open(pdf_path)
                if len(doc) == 0:
                    app_logger.error("[DEBUG] PDF vide ou corrompu")
                    doc.close()
                    return None
                page = doc[0]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom pour meilleure qualité
                pix.save(image_path)
                doc.close()
                app_logger.info(f"[DEBUG] ✓ Conversion réussie avec PyMuPDF: {image_path}")
                return image_path
            except ImportError as e:
                app_logger.warning(f"[DEBUG] PyMuPDF non disponible: {e}")
            except Exception as e:
                app_logger.error(f"[DEBUG] Erreur avec PyMuPDF: {e}", exc_info=True)
            
            app_logger.error("[DEBUG] Aucune bibliothèque PDF disponible (pdf2image ou PyMuPDF)")
            return None
        except Exception as e:
            app_logger.error(f"[DEBUG] Erreur générale lors de la conversion PDF: {e}", exc_info=True)
            return None
    
    def _extract_text(self, filepath, method='auto'):
        """Extrait le texte d'une image"""
        try:
            # Si méthode auto, essayer Ollama puis Tesseract
            if method == 'auto':
                if self.ollama_available:
                    method = 'ollama'
                else:
                    method = 'tesseract'
            
            if method == 'ollama' and self.ollama_available:
                return self._extract_with_ollama(filepath)
            elif method == 'tesseract':
                return self._extract_with_tesseract(filepath)
            else:
                # Fallback sur Tesseract
                return self._extract_with_tesseract(filepath)
        except Exception as e:
            app_logger.error(f"Erreur lors de l'extraction: {e}")
            return None
    
    def _extract_with_ollama(self, image_path):
        """Extrait le texte avec Ollama"""
        try:
            import requests
            import base64
            
            # Lire l'image et la convertir en base64
            with open(image_path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')
            
            # Utiliser le premier modèle disponible
            model = self.ollama_models[0] if self.ollama_models else "llama3.2"
            
            # Appel à Ollama pour extraction OCR
            prompt = """Extrait toutes les informations textuelles de cette image de facture. 
            Retourne uniquement le texte brut, sans formatage."""
            
            response = requests.post(
                f"{self.ollama_base_url}/api/generate",
                json={
                    "model": model,
                    "prompt": prompt,
                    "images": [image_data],
                    "stream": False
                },
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                return result.get('response', '')
            else:
                app_logger.warning(f"Ollama a retourné une erreur: {response.status_code}")
                return None
        except Exception as e:
            app_logger.error(f"Erreur avec Ollama: {e}")
            return None
    
    def _extract_with_tesseract(self, image_path):
        """Extrait le texte avec Tesseract OCR"""
        try:
            import pytesseract
            from PIL import Image
            
            # Tesseract path (peut être configuré)
            tesseract_path = os.environ.get('TESSERACT_PATH', 'tesseract')
            if tesseract_path and os.path.exists(tesseract_path):
                pytesseract.pytesseract.tesseract_cmd = tesseract_path
            
            image = Image.open(image_path)
            text = pytesseract.image_to_string(image, lang='fra+eng')
            return text
        except ImportError:
            app_logger.error("pytesseract non installé. Installez avec: pip install pytesseract")
            return None
        except Exception as e:
            app_logger.error(f"Erreur avec Tesseract: {e}")
            return None
    
    def _parse_invoice_data(self, text):
        """Parse les données de facture depuis le texte extrait"""
        import re
        
        data = {
            "date": "",
            "numero_facture": "",
            "fournisseur": "",
            "total_ht": "",
            "total_ttc": "",
            "tva": "",
            "siret": "",
            "siren": "",
            "adresse_fournisseur": "",
            "code_postal": "",
            "ville": ""
        }
        
        if not text:
            return data
        
        # Extraire la date (format JJ/MM/AAAA ou DD/MM/YYYY)
        date_pattern = r'\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b'
        dates = re.findall(date_pattern, text)
        if dates:
            data["date"] = dates[0]
        
        # Extraire le numéro de facture
        facture_pattern = r'(?:facture|invoice|n[°o]|num[ée]ro)[\s:]*([A-Z0-9\-]+)'
        facture_match = re.search(facture_pattern, text, re.IGNORECASE)
        if facture_match:
            data["numero_facture"] = facture_match.group(1)
        
        # Extraire les montants
        montant_pattern = r'(?:total|montant|amount)[\s:]*([\d\s,\.]+)'
        montants = re.findall(montant_pattern, text, re.IGNORECASE)
        if montants:
            try:
                montant = float(montants[-1].replace(',', '.').replace(' ', ''))
                data["total_ttc"] = montant
            except:
                pass
        
        # Extraire SIRET (14 chiffres)
        siret_pattern = r'\b(\d{14})\b'
        siret_match = re.search(siret_pattern, text)
        if siret_match:
            siret = siret_match.group(1)
            data["siret"] = siret
            data["siren"] = siret[:9]  # SIREN = 9 premiers chiffres
        
        # Extraire le code postal (5 chiffres)
        cp_pattern = r'\b(\d{5})\b'
        cp_match = re.search(cp_pattern, text)
        if cp_match:
            data["code_postal"] = cp_match.group(1)
        
        return data
    
    def _enrich_with_siret(self, data):
        """Enrichit les données avec SIRET via API"""
        # Cette fonction peut être implémentée plus tard avec une vraie API
        return data
    
    def export_to_excel(self, invoice_data, excel_path=None):
        """Exporte les données vers Excel"""
        if not EXCEL_AVAILABLE:
            try:
                import openpyxl
            except ImportError:
                return {
                    'success': False,
                    'error': 'Excel non disponible (openpyxl non installé)'
                }
        
        excel_path = excel_path or self.excel_file_path
        
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.styles.numbers import FORMAT_NUMBER_00
            from openpyxl.utils import get_column_letter
            
            # Créer le dossier si nécessaire
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            
            # Charger ou créer le fichier
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Factures"
                
                # Créer les en-têtes
                headers = ["Date Facture", "Numéro Facture", "Fournisseur", "SIREN", "SIRET",
                          "Adresse", "Code Postal", "Ville", "ID Client", 
                          "Total HT", "TVA", "Total TTC", "Fichier PDF", "Date Extraction"]
                ws.append(headers)
                
                # Formater les en-têtes
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
            
            # Préparer les données
            row_data = [
                invoice_data.get("date", ""),
                invoice_data.get("numero_facture", ""),
                invoice_data.get("fournisseur", ""),
                invoice_data.get("siren", ""),
                invoice_data.get("siret", ""),
                invoice_data.get("adresse_fournisseur", ""),
                invoice_data.get("code_postal", ""),
                invoice_data.get("ville", ""),
                invoice_data.get("id_client", ""),
                invoice_data.get("total_ht", ""),
                invoice_data.get("tva", ""),
                invoice_data.get("total_ttc", ""),
                invoice_data.get("fichier_pdf", ""),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
            
            ws.append(row_data)
            
            # Formater les colonnes numériques
            last_row = ws.max_row
            for col_letter in ['J', 'K', 'L']:  # Total HT, TVA, Total TTC
                cell = ws[f"{col_letter}{last_row}"]
                if cell.value:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = FORMAT_NUMBER_00
                    except:
                        pass
            
            # Sauvegarder
            wb.save(excel_path)
            
            return {
                'success': True,
                'filepath': excel_path,
                'message': 'Données exportées avec succès'
            }
        
        except Exception as e:
            app_logger.error(f"Erreur lors de l'export Excel: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }
